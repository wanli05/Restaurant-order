from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def add_header(ws, row_idx, headers):
    fill = PatternFill("solid", fgColor="DCE6F1")
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=name)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def add_rows(ws, start_row, rows):
    for i, row in enumerate(rows, start=start_row):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if j in (3, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "测试验收记录"

    ws["A1"] = "门店系统测试验收记录表"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "测试日期："
    ws["D3"] = "测试门店："
    ws["A4"] = "测试人员："
    ws["D4"] = "系统版本："
    ws["A5"] = "测试环境：本地 / 测试服 / 其他"
    ws["A6"] = "结果标记：PASS（通过） / FAIL（失败） / N/A（不适用）"

    headers = ["编号", "测试项", "结果", "问题描述（失败时填写）", "测试人", "时间"]
    add_header(ws, 8, headers)

    rows = [
        ("A-01", "点单页可打开", "", "", "", ""),
        ("A-02", "后台页可打开", "", "", "", ""),
        ("A-03", "厨房页可打开", "", "", "", ""),
        ("A-04", "财务页可打开", "", "", "", ""),
        ("A-05", "后台账号可正常登录", "", "", "", ""),
        ("B-01", "默认语言为日语", "", "", "", ""),
        ("B-02", "语言切换（中/日/英）正常", "", "", "", ""),
        ("B-03", "仅选桌号不能下单（人数必填）", "", "", "", ""),
        ("B-04", "仅填人数不能下单（桌号必填）", "", "", "", ""),
        ("B-05", "桌号+人数+菜品后可下单成功", "", "", "", ""),
        ("B-06", "下单成功后购物车清空", "", "", "", ""),
        ("B-07", "我的订单显示正确", "", "", "", ""),
        ("C-01", "厨房页能看到新订单", "", "", "", ""),
        ("C-02", "厨房标记出餐后状态更新", "", "", "", ""),
        ("C-03", "后台可看到结账请求", "", "", "", ""),
        ("C-04", "后台现金收款成功", "", "", "", ""),
        ("C-05", "后台PayPay收款成功", "", "", "", ""),
        ("C-06", "后台支付宝收款成功", "", "", "", ""),
        ("D-01", "历史订单显示指定表头", "", "", "", ""),
        ("D-02", "不显示“已结账/已归档”列", "", "", "", ""),
        ("D-03", "用餐人数无“人数：”前缀", "", "", "", ""),
        ("D-04", "最新订单排序在前", "", "", "", ""),
        ("E-01", "今日营业额正确更新", "", "", "", ""),
        ("E-02", "现金/PayPay/支付宝分项金额正确", "", "", "", ""),
        ("E-03", "客数正确（人数总和）", "", "", "", ""),
        ("E-04", "客单价正确（营业额/成交单数）", "", "", "", ""),
        ("E-05", "中日切换文案正常", "", "", "", ""),
        ("F-01", "不填日期默认导出当天", "", "", "", ""),
        ("F-02", "指定单日导出成功", "", "", "", ""),
        ("F-03", "指定起止日期区间导出成功", "", "", "", ""),
        ("F-04", "开始日期晚于结束日期可拦截", "", "", "", ""),
        ("F-05", "CSV 含客数/客单价新增字段", "", "", "", ""),
        ("G-01", "营业中可执行封账", "", "", "", ""),
        ("G-02", "停止营业时封账按钮禁用", "", "", "", ""),
        ("G-03", "封账前校验弹窗逻辑正确", "", "", "", ""),
        ("G-04", "封账后状态与数据符合预期", "", "", "", ""),
        ("H-01", "网络断开后恢复可继续使用", "", "", "", ""),
        ("H-02", "连点按钮不会重复提交异常订单", "", "", "", ""),
        ("H-03", "刷新页面后数据状态一致", "", "", "", ""),
        ("H-04", "两台设备同桌操作无明显错乱", "", "", "", ""),
        ("I-01", "核心流程可跑通", "", "", "", ""),
        ("I-02", "新增功能全部通过", "", "", "", ""),
        ("I-03", "无阻塞性问题", "", "", "", ""),
    ]
    add_rows(ws, 9, rows)

    end_row = 9 + len(rows) + 1
    ws.cell(row=end_row, column=1, value="本次通过项数量：")
    ws.cell(row=end_row + 1, column=1, value="失败项数量：")
    ws.cell(row=end_row + 2, column=1, value="是否可进入门店试运行：是 / 否")
    ws.cell(row=end_row + 3, column=1, value="结论说明：")
    ws.cell(row=end_row + 4, column=1, value="负责人签字：")
    ws.cell(row=end_row + 5, column=1, value="日期：")

    widths = {
        "A": 10,
        "B": 46,
        "C": 14,
        "D": 46,
        "E": 14,
        "F": 20,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for row in ws.iter_rows(min_row=8, max_row=end_row + 5, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    output = Path("docs") / "store-system-test-checklist.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Generated: {output}")
    print(f"Generated at: {datetime.now().isoformat(timespec='seconds')}")


if __name__ == "__main__":
    main()
